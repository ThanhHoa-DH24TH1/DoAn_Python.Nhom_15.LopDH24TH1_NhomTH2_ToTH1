import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
from datetime import datetime
from dao.student_dao import StudentDAO
from dao.room_dao import RoomDAO
from dao.invoice_dao import InvoiceDAO
from dao.contract_dao import ContractDAO
from models.student import Student
from utils.validator import Validator
import openpyxl
from openpyxl.styles import Font

class StudentManagementForm:
    def __init__(self, parent):
        self.window = tk.Toplevel(parent)
        self.window.title("Quản lý Sinh viên")
        self.window.geometry("1200x700")
        self.window.state('zoomed')
        
        self.student_dao = StudentDAO()
        self.room_dao = RoomDAO()
        self.invoice_dao = InvoiceDAO()
        self.contract_dao = ContractDAO()
        
        self.selected_student = None
        
        self.create_widgets()
        self.load_students()
    
    def create_widgets(self):
        """Tạo giao diện"""
        # Title
        title_frame = tk.Frame(self.window, bg='#2196F3', height=60)
        title_frame.pack(fill='x')
        title_frame.pack_propagate(False)
        
        tk.Label(
            title_frame,
            text="QUẢN LÝ SINH VIÊN",
            font=('Arial', 16, 'bold'),
            bg='#2196F3',
            fg='white'
        ).pack(pady=15)
        
        # Search frame
        search_frame = tk.Frame(self.window, bg='white')
        search_frame.pack(fill='x', padx=10, pady=10)
        
        tk.Label(
            search_frame,
            text="Tìm kiếm:",
            font=('Arial', 10),
            bg='white'
        ).pack(side='left', padx=5)
        
        self.search_entry = tk.Entry(search_frame, font=('Arial', 10), width=30)
        self.search_entry.pack(side='left', padx=5)
        self.search_entry.bind('<KeyRelease>', lambda e: self.search_students())
        
        tk.Label(
            search_frame,
            text="Khoa:",
            font=('Arial', 10),
            bg='white'
        ).pack(side='left', padx=(20, 5))
        
        self.faculty_combo = ttk.Combobox(
            search_frame,
            font=('Arial', 10),
            width=20,
            state='readonly'
        )
        self.faculty_combo.pack(side='left', padx=5)
        self.faculty_combo.bind('<<ComboboxSelected>>', lambda e: self.search_students())
        
        tk.Button(
            search_frame,
            text="🔄 Làm mới",
            font=('Arial', 10),
            bg='#4CAF50',
            fg='white',
            cursor='hand2',
            command=self.load_students
        ).pack(side='left', padx=5)
        
        # Main content
        content_frame = tk.Frame(self.window)
        content_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Treeview
        tree_frame = tk.Frame(content_frame)
        tree_frame.pack(side='left', fill='both', expand=True)
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        # Treeview
        columns = ('STT', 'MSSV', 'Họ tên', 'Ngày sinh', 'Giới tính', 
                  'SĐT', 'Khoa', 'Lớp', 'Phòng', 'Trạng thái')
        
        self.tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show='headings',
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set
        )
        
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        
        # Column headings
        widths = [50, 100, 200, 100, 80, 120, 150, 100, 80, 120]
        for col, width in zip(columns, widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor='center')
        
        self.tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')
        hsb.pack(side='bottom', fill='x')
        
        # Double click to view details
        self.tree.bind('<Double-1>', self.on_double_click)
        
        # Button frame
        btn_frame = tk.Frame(content_frame, width=150)
        btn_frame.pack(side='right', fill='y', padx=(10, 0))
        btn_frame.pack_propagate(False)
        
        buttons = [
            ("➕ Thêm", self.add_student, '#4CAF50'),
            ("✏️ Sửa", self.edit_student, '#2196F3'),
            ("🗑️ Xóa", self.delete_student, '#f44336'),
            ("👁️ Chi tiết", self.view_details, '#FF9800'),
            ("📄 Xuất Excel", self.export_excel, '#9C27B0')
        ]
        
        for text, cmd, color in buttons:
            btn = tk.Button(
                btn_frame,
                text=text,
                font=('Arial', 10),
                bg=color,
                fg='white',
                cursor='hand2',
                width=15,
                command=cmd
            )
            btn.pack(pady=5, fill='x')
        
        # Load faculties
        self.load_faculties()
    
    def load_faculties(self):
        """Load danh sách khoa"""
        faculties = self.student_dao.get_faculties()
        self.faculty_combo['values'] = ['Tất cả'] + faculties
        self.faculty_combo.current(0)
    
    def load_students(self):
        """Load danh sách sinh viên"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        students = self.student_dao.get_all_students()
        
        for idx, student in enumerate(students, 1):
            # student: (StudentID, StudentCode, FullName, DOB, Gender, Phone, 
            #           Email, IDCard, Address, Faculty, Major, Class, Status, UserID, RoomNumber)
            values = (
                idx,
                student[1],  # StudentCode
                student[2],  # FullName
                student[3].strftime('%d/%m/%Y') if student[3] else '',
                student[4],  # Gender
                student[5] or '',  # Phone
                student[9] or '',  # Faculty
                student[11] or '',  # Class
                student[14] or '',  # RoomNumber
                student[12]  # Status
            )
            self.tree.insert('', 'end', values=values, tags=(student[0],))
    
    def search_students(self):
        """Tìm kiếm sinh viên"""
        keyword = self.search_entry.get().strip()
        faculty = self.faculty_combo.get()
        faculty = None if faculty == 'Tất cả' else faculty
        
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        if not keyword:
            self.load_students()
            return
        
        students = self.student_dao.search_students(keyword, faculty)
        
        for idx, student in enumerate(students, 1):
            values = (
                idx,
                student[1],
                student[2],
                student[3].strftime('%d/%m/%Y') if student[3] else '',
                student[4],
                student[5] or '',
                student[9] or '',
                student[11] or '',
                student[14] or '',
                student[12]
            )
            self.tree.insert('', 'end', values=values, tags=(student[0],))
    
    def on_double_click(self, event):
        """Xử lý double click"""
        self.view_details()
    
    def add_student(self):
        """Thêm sinh viên"""
        StudentFormDialog(self.window, None, self.on_save_success)
    
    def edit_student(self):
        """Sửa sinh viên"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn sinh viên cần sửa!")
            return
        
        student_id = self.tree.item(selected[0])['tags'][0]
        student_data = self.student_dao.get_student_by_id(student_id)
        
        if student_data:
            StudentFormDialog(self.window, student_data, self.on_save_success)
    
    def delete_student(self):
        """Xóa sinh viên"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn sinh viên cần xóa!")
            return

        if not messagebox.askyesno("Xác nhận", 
            "Bạn có chắc muốn xóa sinh viên này?\n\n"
            "CẢNH BÁO: Tất cả hợp đồng và hóa đơn liên quan cũng sẽ bị xóa vĩnh viễn!"): # Thêm cảnh báo
            return

        try:
            student_id = self.tree.item(selected[0])['tags'][0]

            # ===== BƯỚC 1: XÓA PAYMENTS (MỚI THÊM) =====
            print(f"Đang xóa payments cho StudentID: {student_id}")
            deleted_payments = self.invoice_dao.delete_payments_by_student(student_id)
            print(f"Số payments đã xóa: {deleted_payments}") 

            # ===== BƯỚC 2: XÓA HÓA ĐƠN =====
            print(f"Đang xóa hóa đơn cho StudentID: {student_id}")
            deleted_invoices = self.invoice_dao.delete_invovice_by_student(student_id)
            print(f"Số hóa đơn đã xóa: {deleted_invoices}")

            # ===== BƯỚC 3: XÓA HỢP ĐỒNG =====
            print(f"Đang xóa hợp đồng cho StudentID: {student_id}")
            deleted_contracts = self.contract_dao.delete_student_contracts(student_id)
            print(f"Số hợp đồng đã xóa: {deleted_contracts}") 

            # ===== BƯỚC 4: XÓA SINH VIÊN =====
            print(f"Đang xóa sinh viên StudentID: {student_id}")
            if self.student_dao.delete_student(student_id):
                messagebox.showinfo("Thành công", "Đã xóa sinh viên và các dữ liệu liên quan!")
                self.load_students() 
            else:
                messagebox.showerror("Lỗi", "Không thể xóa sinh viên (sau khi đã xóa dữ liệu liên quan).")

        except Exception as e:
            messagebox.showerror("Lỗi Xóa", f"Đã xảy ra lỗi trong quá trình xóa:\n{e}")
            import traceback
            traceback.print_exc()

    def view_details(self):
        """Xem chi tiết"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn sinh viên!")
            return
        
        student_id = self.tree.item(selected[0])['tags'][0]
        student_data = self.student_dao.get_student_by_id(student_id)
        
        if student_data:
            detail_text = f"""
            THÔNG TIN CHI TIẾT SINH VIÊN
            
            MSSV: {student_data[1]}
            Họ và tên: {student_data[2]}
            Ngày sinh: {student_data[3].strftime('%d/%m/%Y')}
            Giới tính: {student_data[4]}
            CMND/CCCD: {student_data[7]}
            Số điện thoại: {student_data[5] or 'Chưa có'}
            Email: {student_data[6] or 'Chưa có'}
            Địa chỉ: {student_data[8] or 'Chưa có'}
            
            Khoa: {student_data[9]}
            Chuyên ngành: {student_data[10] or 'Chưa có'}
            Lớp: {student_data[11]}
            Trạng thái: {student_data[12]}
            """
            messagebox.showinfo("Chi tiết sinh viên", detail_text)
    
    def export_excel(self):
        """Xuất Excel"""
        file_path = filedialog.asksaveasfilename(
            title="Lưu file Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        # 2. Nếu người dùng không chọn (nhấn Cancel) thì dừng lại
        if not file_path:
            return

        try:
            # 3. Lấy dữ liệu từ Treeview
            # Lấy tiêu đề cột
            headers = list(self.tree['columns'])
            
            # Lấy dữ liệu từng hàng
            data = []
            for item_id in self.tree.get_children():
                row_values = self.tree.item(item_id)['values']
                data.append(list(row_values)) # Chuyển tuple sang list

            # 4. Tạo file Excel mới
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "DanhSachSinhVien"
            
            # 5. Ghi tiêu đề (in đậm)
            ws.append(headers)
            for cell in ws[1]: # ws[1] là hàng đầu tiên
                cell.font = Font(bold=True)
            
            # 6. Ghi dữ liệu từng hàng
            for row in data:
                ws.append(row)
                
            # 7. (Tùy chọn) Tự động điều chỉnh độ rộng cột
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Lấy tên cột (A, B, C...)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # 8. Lưu file 💾
            wb.save(file_path)
            
            messagebox.showinfo("Thành công", f"Đã xuất dữ liệu ra file:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu file Excel:\n{e}")
    
    def on_save_success(self):
        """Callback khi lưu thành công"""
        self.load_students()
class StudentFormDialog:
    def __init__(self, parent, student_data, callback):
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Thông tin sinh viên")
        self.dialog.geometry("600x700")
        self.dialog.resizable(False, False)
        self.dialog.grab_set()
        
        self.student_data = student_data
        self.callback = callback
        self.student_dao = StudentDAO()
        self.room_dao = RoomDAO()
        
        self.create_form()
        
        if student_data:
            self.load_data()
    
    def create_form(self):
        """Tạo form"""
        main_container = tk.Frame(self.dialog)
        main_container.pack(fill='both', expand=True)
        # Title
        title = "CẬP NHẬT SINH VIÊN" if self.student_data else "THÊM SINH VIÊN MỚI"
        tk.Label(
            self.dialog,
            text=title,
            font=('Arial', 14, 'bold'),
            fg='#2196F3'
        ).pack(pady=15)
        
        # Form frame
        form_frame = tk.Frame(self.dialog)
        form_frame.pack(fill='both', expand=True, padx=20)
        
        # Fields
        fields = [
            ("MSSV (*)", "mssv"),
            ("Họ và tên (*)", "fullname"),
            ("Ngày sinh (*)", "dob"),
            ("Giới tính (*)", "gender"),
            ("CMND/CCCD (*)", "idcard"),
            ("Số điện thoại", "phone"),
            ("Email", "email"),
            ("Địa chỉ", "address"),
            ("Khoa (*)", "faculty"),
            ("Chuyên ngành", "major"),
            ("Lớp (*)", "class"),
        ]
        
        self.entries = {}
        
        for idx, (label, key) in enumerate(fields):
            tk.Label(
                form_frame,
                text=label,
                font=('Arial', 10),
                anchor='w'
            ).grid(row=idx, column=0, sticky='w', pady=8)
            
            if key == 'dob':
                self.entries[key] = DateEntry(
                    form_frame,
                    font=('Arial', 10),
                    width=35,
                    date_pattern='dd/mm/yyyy'
                )
            elif key == 'gender':
                gender_frame = tk.Frame(form_frame)
                self.entries[key] = tk.StringVar(value='Nam')
                tk.Radiobutton(
                    gender_frame,
                    text='Nam',
                    variable=self.entries[key],
                    value='Nam',
                    font=('Arial', 10)
                ).pack(side='left', padx=10)
                tk.Radiobutton(
                    gender_frame,
                    text='Nữ',
                    variable=self.entries[key],
                    value='Nữ',
                    font=('Arial', 10)
                ).pack(side='left')
                gender_frame.grid(row=idx, column=1, sticky='w', pady=8)
                continue
            elif key == 'address':
                self.entries[key] = tk.Text(form_frame, font=('Arial', 10), width=37, height=3)
            else:
                self.entries[key] = tk.Entry(form_frame, font=('Arial', 10), width=37)
            
            self.entries[key].grid(row=idx, column=1, pady=8)
        #tạo nút lưu hủy
        btn_frame = tk.Frame(self.dialog)
        btn_frame.pack(pady=20)
       
        tk.Button(
            btn_frame,
            text="💾 Lưu",
            font=('Arial', 10, 'bold'),
            bg='#4CAF50',
            fg='white',
            width=12,
            cursor='hand2',
            command=self.save
        ).pack(side='left', padx=5)
        
        tk.Button(
            btn_frame,
            text="❌ Hủy",
            font=('Arial', 10),
            bg='#f44336',
            fg='white',
            width=12,
            cursor='hand2',
            command=self.dialog.destroy
        ).pack(side='left', padx=5)
    
        # Phòng (nếu thêm mới)
        row_idx = len(fields)
        if not self.student_data:
            tk.Label(
                form_frame,
                text="Phòng:",
                font=('Arial', 10),
                anchor='w'
            ).grid(row=row_idx, column=0, sticky='w', pady=8)
            
            self.room_combo = ttk.Combobox(
                form_frame,
                font=('Arial', 10),
                width=35,
                state='readonly'
            )
            self.room_combo.grid(row=row_idx, column=1, pady=8)
            self.load_available_rooms()
            row_idx += 1
       
    def load_available_rooms(self):
        """Load phòng còn chỗ"""
        rooms = self.room_dao.get_available_rooms()
        room_list = ['Chưa chọn'] + [f"{r[1]} - {r[2]}{r[3]} (Còn {r[5]-r[6]} chỗ)" for r in rooms]
        self.room_combo['values'] = room_list
        self.room_combo.current(0)
    
    def load_data(self):
        """Load dữ liệu sinh viên"""
        self.entries['mssv'].insert(0, self.student_data[1])
        self.entries['mssv'].config(state='disabled')  # Không cho sửa MSSV
        
        self.entries['fullname'].insert(0, self.student_data[2])
        self.entries['dob'].set_date(self.student_data[3])
        self.entries['gender'].set(self.student_data[4])
        self.entries['idcard'].insert(0, self.student_data[7])
        
        if self.student_data[5]:
            self.entries['phone'].insert(0, self.student_data[5])
        if self.student_data[6]:
            self.entries['email'].insert(0, self.student_data[6])
        if self.student_data[8]:
            self.entries['address'].insert('1.0', self.student_data[8])
        
        self.entries['faculty'].insert(0, self.student_data[9])
        if self.student_data[10]:
            self.entries['major'].insert(0, self.student_data[10])
        self.entries['class'].insert(0, self.student_data[11])
    
    def validate(self):
        """Validate dữ liệu"""
        # MSSV
        if not self.student_data:  # Chỉ validate khi thêm mới
            mssv = self.entries['mssv'].get().strip()
            valid, msg = Validator.validate_student_code(mssv)
            if not valid:
                messagebox.showwarning("Cảnh báo", msg)
                self.entries['mssv'].focus()
                return False
        
        # Họ tên
        fullname = self.entries['fullname'].get().strip()
        valid, msg = Validator.validate_full_name(fullname)
        if not valid:
            messagebox.showwarning("Cảnh báo", msg)
            self.entries['fullname'].focus()
            return False
        
        # CMND
        idcard = self.entries['idcard'].get().strip()
        valid, msg = Validator.validate_id_card(idcard)
        if not valid:
            messagebox.showwarning("Cảnh báo", msg)
            self.entries['idcard'].focus()
            return False
        
        # SĐT (nếu có)
        phone = self.entries['phone'].get().strip()
        if phone:
            valid, msg = Validator.validate_phone(phone)
            if not valid:
                messagebox.showwarning("Cảnh báo", msg)
                self.entries['phone'].focus()
                return False
        
        # Email (nếu có)
        email = self.entries['email'].get().strip()
        if email:
            valid, msg = Validator.validate_email(email)
            if not valid:
                messagebox.showwarning("Cảnh báo", msg)
                self.entries['email'].focus()
                return False
        
        # Khoa
        faculty = self.entries['faculty'].get().strip()
        if not faculty:
            messagebox.showwarning("Cảnh báo", "Khoa không được để trống!")
            self.entries['faculty'].focus()
            return False
        
        # Lớp
        class_name = self.entries['class'].get().strip()
        if not class_name:
            messagebox.showwarning("Cảnh báo", "Lớp không được để trống!")
            self.entries['class'].focus()
            return False
        
        return True
    
    def save(self):
        """Lưu sinh viên"""
        if not self.validate():
            return
        
        # Thu thập dữ liệu
        student = Student()
        
        if self.student_data:
            student.student_id = self.student_data[0]
            student.student_code = self.student_data[1]
        else:
            student.student_code = self.entries['mssv'].get().strip()
        
        student.full_name = self.entries['fullname'].get().strip()
        student.date_of_birth = self.entries['dob'].get_date()
        student.gender = self.entries['gender'].get()
        student.id_card = self.entries['idcard'].get().strip()
        student.phone = self.entries['phone'].get().strip()
        student.email = self.entries['email'].get().strip()
        student.address = self.entries['address'].get('1.0', 'end').strip()
        student.faculty = self.entries['faculty'].get().strip()
        student.major = self.entries['major'].get().strip()
        student.class_name = self.entries['class'].get().strip()
        student.status = 'Đang ở'
        
        try:
            if self.student_data:
                # Cập nhật
                if self.student_dao.update_student(student):
                    messagebox.showinfo("Thành công", "Đã cập nhật sinh viên!")
                    self.callback()
                    self.dialog.destroy()
                else:
                    messagebox.showerror("Lỗi", "Không thể cập nhật sinh viên!")
            else:
                # Thêm mới
                if self.student_dao.add_student(student):
                    messagebox.showinfo("Thành công", "Đã thêm sinh viên!")
                    
                    # Nếu chọn phòng, tạo hợp đồng
                    if hasattr(self, 'room_combo') and self.room_combo.current() > 0:
                        self.assign_room(student.student_code)
                    
                    self.callback()
                    self.dialog.destroy()
                else:
                    messagebox.showerror("Lỗi", "Không thể thêm sinh viên! MSSV có thể đã tồn tại.")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi lưu: {e}")
    
    def assign_room(self, student_code):
        """Phân phòng cho sinh viên"""
        try:
            from dao.contract_dao import ContractDAO
            from utils.date_utils import DateUtils
            
            # Lấy StudentID
            student = self.student_dao.get_student_by_code(student_code)
            if not student:
                return
            
            student_id = student[0]
            
            # Lấy RoomID từ combo
            room_text = self.room_combo.get()
            if room_text == 'Chưa chọn':
                return
            
            room_number = room_text.split(' - ')[0]
            rooms = self.room_dao.search_rooms()
            room_id = None
            price = 0
            
            for r in rooms:
                if r[1] == room_number:
                    room_id = r[0]
                    price = r[7]
                    break
            
            if not room_id:
                return
            
            # Tạo hợp đồng
            contract_dao = ContractDAO()
            start_date = DateUtils.get_current_date()
            end_date = DateUtils.add_months(start_date, 10)  # 10 tháng
            
            success = contract_dao.add_contract(
                student_id, room_id, start_date, end_date,
                price, 500000, 'Hợp đồng tự động khi thêm SV'
            )
            
            if success:
                # Cập nhật số người ở
                self.room_dao.update_occupancy(room_id)
                messagebox.showinfo("Thành công", f"Đã phân phòng {room_number} cho sinh viên!")
        
        except Exception as e:
            print(f"Lỗi phân phòng: {e}")