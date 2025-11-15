import tkinter as tk
import openpyxl
from openpyxl.styles import Font
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from dao.invoice_dao import InvoiceDAO
from dao.contract_dao import ContractDAO
from config import DEFAULT_PRICES
from decimal import Decimal 

class PaymentManagementForm:
    def __init__(self, parent):
        self.window = tk.Toplevel(parent)
        self.window.title("Quản lý Thanh toán")
        self.window.geometry("1400x700")
        self.window.state('zoomed')
        
        self.invoice_dao = InvoiceDAO()
        self.contract_dao = ContractDAO()
        
        self.selected_invoice = None
        
        self.create_widgets()
        self.load_invoices()
    
    def create_widgets(self):
        """Tạo giao diện"""
        # Title
        title_frame = tk.Frame(self.window, bg='#FF9800', height=60)
        title_frame.pack(fill='x')
        title_frame.pack_propagate(False)
        
        tk.Label(
            title_frame,
            text="QUẢN LÝ THANH TOÁN",
            font=('Arial', 16, 'bold'),
            bg='#FF9800',
            fg='white'
        ).pack(pady=15)
        
        # Filter frame
        filter_frame = tk.Frame(self.window, bg='white')
        filter_frame.pack(fill='x', padx=10, pady=10)
        
        tk.Label(filter_frame, text="Tháng:", bg='white').pack(side='left', padx=5)
        self.month_combo = ttk.Combobox(filter_frame, width=10, state='readonly')
        self.month_combo['values'] = self.get_months()
        self.month_combo.current(0)
        self.month_combo.pack(side='left', padx=5)
        
        tk.Label(filter_frame, text="Trạng thái:", bg='white').pack(side='left', padx=5)
        self.status_combo = ttk.Combobox(filter_frame, width=15, state='readonly')
        self.status_combo['values'] = ['Tất cả', 'Chưa thanh toán', 'Đã thanh toán']
        self.status_combo.current(0)
        self.status_combo.pack(side='left', padx=5)
        
        tk.Label(filter_frame, text="Tìm:", bg='white').pack(side='left', padx=5)
        self.search_entry = tk.Entry(filter_frame, width=20)
        self.search_entry.pack(side='left', padx=5)
        
        tk.Button(
            filter_frame,
            text="🔍 Tìm",
            bg='#2196F3',
            fg='white',
            command=self.search_invoices
        ).pack(side='left', padx=5)
        
        tk.Button(
            filter_frame,
            text="🔄 Làm mới",
            bg='#4CAF50',
            fg='white',
            command=self.load_invoices
        ).pack(side='left', padx=5)
        
        # Main content
        content_frame = tk.Frame(self.window)
        content_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Left: Treeview
        left_frame = tk.Frame(content_frame)
        left_frame.pack(side='left', fill='both', expand=True)
        
        columns = ('STT', 'Mã HĐ', 'MSSV', 'Họ tên', 'Phòng', 'Tháng', 
                  'Tổng tiền', 'Đã trả', 'Còn nợ', 'Trạng thái')
        
        self.tree = ttk.Treeview(left_frame, columns=columns, show='headings')
        
        widths = [40, 70, 100, 150, 80, 80, 100, 100, 100, 120]
        for col, width in zip(columns, widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor='center')
        
        vsb = ttk.Scrollbar(left_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        
        self.tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')
        
        self.tree.bind('<<TreeviewSelect>>', self.on_select)
        
        # Right: Detail panel
        right_frame = tk.LabelFrame(
            content_frame,
            text="Chi tiết hóa đơn",
            font=('Arial', 11, 'bold'),
            width=400
        )
        right_frame.pack(side='right', fill='y', padx=(10, 0))
        right_frame.pack_propagate(False)
        
        self.detail_text = tk.Text(
            right_frame,
            font=('Arial', 10),
            wrap='word',
            state='disabled'
        )
        self.detail_text.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Bottom: Buttons
        btn_frame = tk.Frame(self.window, bg='white')
        btn_frame.pack(fill='x', padx=10, pady=10)
        
        buttons = [
            ("📝 Tạo HĐ tháng", self.create_monthly_invoices, '#4CAF50'),
            ("💰 Thanh toán", self.record_payment, '#2196F3'),
            ("📊 Xuất công nợ", self.export_debt, '#9C27B0')
        ]
        
        for text, cmd, color in buttons:
            tk.Button(
                btn_frame,
                text=text,
                font=('Arial', 10),
                bg=color,
                fg='white',
                width=20,
                command=cmd
            ).pack(side='left', padx=5)
    
    def get_months(self):
        """Lấy danh sách tháng"""
        months = []
        for i in range(12, -1, -1):
            date = datetime.now()
            month = date.month - i
            year = date.year
            if month <= 0:
                month += 12
                year -= 1
            months.append(f"{year}-{month:02d}")
        return months
    
    def load_invoices(self):
        """Load danh sách hóa đơn"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        month = self.month_combo.get() 
        print(month)
        status = self.status_combo.get()
        print(status)
        status = None if status == 'Tất cả' else status
        
        invoices = self.invoice_dao.get_all_invoices(month, status)
        
        for idx, inv in enumerate(invoices, 1):
            values = (
                idx,
                inv[0],  # InvoiceID
                inv[17],  # StudentCode
                inv[18],  # FullName
                inv[19],  # RoomNumber
                inv[3],  # BillingMonth
                f"{inv[9]:,.0f}",  # TotalAmount
                f"{inv[10]:,.0f}",  # PaidAmount
                f"{inv[11]:,.0f}",  # RemainingAmount
                inv[12]  # Status
            )
            self.tree.insert('', 'end', values=values, tags=(inv[0],))
    
    def search_invoices(self):
        """Tìm kiếm hóa đơn"""
        self.load_invoices()
    
    def on_select(self, event):
        """Khi chọn hóa đơn"""
        selected = self.tree.selection()
        if not selected:
            return
        
        invoice_id = self.tree.item(selected[0])['tags'][0]
        invoice = self.invoice_dao.get_invoice_by_id(invoice_id)
        
        if invoice:
            self.selected_invoice = invoice
            self.display_invoice_detail(invoice)
    
    def display_invoice_detail(self, invoice):
        """Hiển thị chi tiết hóa đơn"""
        detail = f"""
HÓA ĐƠN TIỀN PHÒNG
{'='*40}

Mã hóa đơn: {invoice[0]}
MSSV: {invoice[17]}
Họ tên: {invoice[18]}
Phòng: {invoice[19]}
Tháng: {invoice[3]}
Ngày tạo: {invoice[15].strftime('%d/%m/%Y %H:%M')}

CHI TIẾT CÁC KHOẢN PHÍ:
{'-'*40}
1. Tiền phòng:        {invoice[4]:>15,.0f} đ
2. Tiền điện:         {invoice[5]:>15,.0f} đ
3. Tiền nước:         {invoice[6]:>15,.0f} đ
4. Phí internet:      {invoice[7]:>15,.0f} đ
5. Phí dịch vụ:       {invoice[8]:>15,.0f} đ
{'='*40}
TỔNG CỘNG:            {invoice[9]:>15,.0f} đ
Đã thanh toán:        {invoice[10]:>15,.0f} đ
CÒN NỢ:               {invoice[11]:>15,.0f} đ

Trạng thái: {invoice[12]}
        """
        
        self.detail_text.config(state='normal')
        self.detail_text.delete('1.0', 'end')
        self.detail_text.insert('1.0', detail)
        self.detail_text.config(state='disabled')
    
    def create_monthly_invoices(self):
        """Tạo hóa đơn hàng tháng"""
        CreateInvoiceDialog(self.window, self.contract_dao, 
                          self.invoice_dao, self.load_invoices)
    
    def record_payment(self):
        """Ghi nhận thanh toán"""
        if not self.selected_invoice:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn hóa đơn!")
            return
        
        if self.selected_invoice[11] <= 0:
            messagebox.showinfo("Thông báo", "Hóa đơn đã thanh toán đầy đủ!")
            return
        
        PaymentDialog(self.window, self.selected_invoice, 
                     self.invoice_dao, self.load_invoices)
    
    def export_debt(self):
        """Xuất danh sách công nợ (hóa đơn còn nợ) ra file Excel."""
        
        # 1. Hỏi người dùng muốn lưu file ở đâu
        file_path = filedialog.asksaveasfilename(
            title="Lưu file Công nợ",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"CongNo_Thang_{self.month_combo.get()}.xlsx" # Tên file gợi ý
        )
        
        # 2. Nếu người dùng không chọn (nhấn Cancel) thì dừng lại
        if not file_path:
            return

        try:
            # 3. Tạo file Excel mới
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "DanhSachCongNo"
            
            # 4. Lấy và ghi tiêu đề (in đậm)
            # Chọn các cột cần xuất
            headers = ['Tháng', 'MSSV', 'Họ tên', 'Phòng', 'Tổng tiền', 'Đã trả', 'Còn nợ']
            ws.append(headers)
            header_font = Font(bold=True)
            for cell in ws[1]: # ws[1] là hàng đầu tiên
                cell.font = header_font
            
            # 5. Lấy dữ liệu từ Treeview và lọc hóa đơn còn nợ
            total_debt_amount = Decimal(0)
            debt_rows = []
            
            # Lấy các cột tương ứng từ Treeview của bạn
            # ('STT', 'Mã HĐ', 'MSSV', 'Họ tên', 'Phòng', 'Tháng', 
            #  'Tổng tiền', 'Đã trả', 'Còn nợ', 'Trạng thái')
            col_indices = {
                'Tháng': 5, 
                'MSSV': 2, 
                'Họ tên': 3, 
                'Phòng': 4, 
                'Tổng tiền': 6, 
                'Đã trả': 7, 
                'Còn nợ': 8
            }
            
            for item_id in self.tree.get_children():
                item_values = self.tree.item(item_id)['values']
                
                try:
                    # Lấy số tiền còn nợ (cột 8) và chuyển lại thành Decimal
                    remaining_str = str(item_values[col_indices['Còn nợ']]).replace(',', '') # Bỏ dấu phẩy
                    remaining_amount = Decimal(remaining_str)
                    
                    # Chỉ thêm vào danh sách nếu còn nợ > 0
                    if remaining_amount > 0:
                        row_data = [
                            item_values[col_indices['Tháng']],
                            item_values[col_indices['MSSV']],
                            item_values[col_indices['Họ tên']],
                            item_values[col_indices['Phòng']],
                            Decimal(str(item_values[col_indices['Tổng tiền']]).replace(',', '')), # Chuyển về Decimal
                            Decimal(str(item_values[col_indices['Đã trả']]).replace(',', '')),   # Chuyển về Decimal
                            remaining_amount
                        ]
                        debt_rows.append(row_data)
                        total_debt_amount += remaining_amount
                        
                except (ValueError, IndexError):
                    print(f"Bỏ qua dòng lỗi dữ liệu: {item_values}") 
                    continue # Bỏ qua nếu có lỗi chuyển đổi số

            # 6. Ghi dữ liệu công nợ vào file
            for row in debt_rows:
                 # Ghi dữ liệu số dưới dạng số để Excel tính toán được
                ws.append(row)
                # Định dạng tiền tệ cho các cột tiền
                for col_letter in ['E', 'F', 'G']: # Cột Tổng tiền, Đã trả, Còn nợ
                    ws[f'{col_letter}{ws.max_row}'].number_format = '#,##0'

            # 7. Thêm dòng tổng cộng ở cuối
            ws.append([]) # Thêm một hàng trống
            total_row_idx = ws.max_row + 1
            total_cell = ws.cell(row=total_row_idx, column=6) # Ghi ở cột "Đã trả" (F)
            total_cell.value = "TỔNG CỘNG NỢ:"
            total_cell.font = Font(bold=True)
            
            total_amount_cell = ws.cell(row=total_row_idx, column=7) # Ghi ở cột "Còn nợ" (G)
            total_amount_cell.value = total_debt_amount
            total_amount_cell.font = Font(bold=True, color="FF0000") # Màu đỏ
            total_amount_cell.number_format = '#,##0'
                
            # 8. Tự động điều chỉnh độ rộng cột
            for col_idx, header in enumerate(headers, 1):
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                max_length = len(header)
                for row_num in range(2, ws.max_row + 1): # Bắt đầu từ hàng 2
                    cell_value = ws.cell(row=row_num, column=col_idx).value
                    if cell_value is not None:
                         # Nếu là số Decimal, định dạng trước khi đo độ dài
                        if isinstance(cell_value, Decimal):
                             cell_display = f"{cell_value:,.0f}" # Thêm dấu phẩy
                        else:
                             cell_display = str(cell_value)
                        max_length = max(max_length, len(cell_display))
                        
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width
            # Điều chỉnh riêng cột Họ tên rộng hơn
            ws.column_dimensions['C'].width = 25 


            # 9. Lưu file 💾
            wb.save(file_path)
            
            messagebox.showinfo("Thành công", f"Đã xuất danh sách công nợ ra file:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu file Excel:\n{e}")
            import traceback
            traceback.print_exc() # In lỗi chi tiết ra terminal


# ============================================
# Dialog tạo hóa đơn tháng
# ============================================

class CreateInvoiceDialog:
    def __init__(self, parent, contract_dao, invoice_dao, callback):
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Tạo hóa đơn tháng")
        self.dialog.geometry("400x400")
        self.dialog.grab_set()
        
        self.contract_dao = contract_dao
        self.invoice_dao = invoice_dao
        self.callback = callback
        
        self.create_widgets()
    
    def create_widgets(self):
        """Tạo form"""
        tk.Label(
            self.dialog,
            text="TẠO HÓA ĐƠN THÁNG",
            font=('Arial', 14, 'bold'),
            fg='#4CAF50'
        ).pack(pady=15)
        
        form_frame = tk.Frame(self.dialog)
        form_frame.pack(padx=20, fill='both', expand=True)
        
        # Tháng/năm
        tk.Label(form_frame, text="Tháng/Năm (*):").grid(row=0, column=0, sticky='w', pady=10)
        self.month_combo = ttk.Combobox(form_frame, width=25, state='readonly')
        self.month_combo['values'] = [datetime.now().strftime('%Y-%m')]
        self.month_combo.current(0)
        self.month_combo.grid(row=0, column=1, pady=10)
        
        # Giá điện
        tk.Label(form_frame, text="Đơn giá điện (đ/kWh):").grid(row=1, column=0, sticky='w', pady=10)
        self.elec_entry = tk.Entry(form_frame, width=27)
        self.elec_entry.insert(0, DEFAULT_PRICES['electricity'])
        self.elec_entry.grid(row=1, column=1, pady=10)
        
        # Giá nước
        tk.Label(form_frame, text="Đơn giá nước (đ/m³):").grid(row=2, column=0, sticky='w', pady=10)
        self.water_entry = tk.Entry(form_frame, width=27)
        self.water_entry.insert(0, DEFAULT_PRICES['water'])
        self.water_entry.grid(row=2, column=1, pady=10)
        
        # Internet
        tk.Label(form_frame, text="Phí internet (đ/tháng):").grid(row=3, column=0, sticky='w', pady=10)
        self.internet_entry = tk.Entry(form_frame, width=27)
        self.internet_entry.insert(0, DEFAULT_PRICES['internet'])
        self.internet_entry.grid(row=3, column=1, pady=10)
        
        # Dịch vụ
        tk.Label(form_frame, text="Phí dịch vụ (đ/tháng):").grid(row=4, column=0, sticky='w', pady=10)
        self.service_entry = tk.Entry(form_frame, width=27)
        self.service_entry.insert(0, DEFAULT_PRICES['cleaning'])
        self.service_entry.grid(row=4, column=1, pady=10)
        
        # Buttons
        btn_frame = tk.Frame(self.dialog)
        btn_frame.pack(pady=20)
        
        tk.Button(
            btn_frame,
            text="✅ Tạo hóa đơn",
            bg='#4CAF50',
            fg='white',
            width=15,
            command=self.create
        ).pack(side='left', padx=5)
        
        tk.Button(
            btn_frame,
            text="❌ Hủy",
            bg='#f44336',
            fg='white',
            width=15,
            command=self.dialog.destroy
        ).pack(side='left', padx=5)
    
    def create(self):
        """Tạo hóa đơn"""
        print("=" * 50)
        print("BẮT ĐẦU TẠO HÓA ĐƠN")
        print("=" * 50)
        try:
            month_str = self.month_combo.get() # Lấy chuỗi ví dụ: "2025-10"
            print(f"Chuỗi tháng/năm: {month_str}")

            try:
                elec_price = float(self.elec_entry.get())
                water_price = float(self.water_entry.get())
                internet_fee = float(self.internet_entry.get())
                service_fee = float(self.service_entry.get())
            except ValueError:
                messagebox.showerror("Lỗi", "Giá dịch vụ không hợp lệ!")
                return
            
            # Kiểm tra tháng đã tạo hóa đơn chưa
            existing = self.invoice_dao.get_all_invoices(month_str)
            print(f"Kiểm tra hóa đơn T{month_str}: Tìm thấy {len(existing)}")
            if existing:
                if not messagebox.askyesno(
                    "Xác nhận", 
                    f"Đã có {len(existing)} hóa đơn cho tháng {month_str}.\nTạo thêm?"
                ):
                    return
            
            # Lấy tất cả hợp đồng đang hiệu lực
            all_contracts = self.contract_dao.get_all_contracts()
            print(f"Tổng số hợp đồng: {len(all_contracts)}")
            if not all_contracts:
                messagebox.showwarning("Cảnh báo", "Không có hợp đồng nào trong hệ thống!")
                return
            
            # Lọc hợp đồng đang hiệu lực
            active_contracts = []
            for contract in all_contracts:
                # contract[7] là Status
                if contract[7] == 'Đang hiệu lực':
                    active_contracts.append(contract)
            print(f"Hợp đồng đang hiệu lực: {len(active_contracts)}")

            if not active_contracts:
                messagebox.showwarning(
                    "Cảnh báo", 
                    f"Không có hợp đồng nào đang hiệu lực!\n\n"
                    f"Tổng số hợp đồng: {len(all_contracts)}\n"
                    f"- Đang hiệu lực: 0\n"
                    f"- Khác: {len(all_contracts)}"
                )
                return
            
            # Confirm
            if not messagebox.askyesno(
                "Xác nhận",
                f"Tạo hóa đơn cho {len(active_contracts)} hợp đồng?\n\n"
                f"Tháng: {month_str}\n"
                f"Giá điện: {elec_price:,.0f} đ/kWh\n"
                f"Giá nước: {water_price:,.0f} đ/m³\n"
                f"Phí internet: {internet_fee:,.0f} đ\n"
                f"Phí dịch vụ: {service_fee:,.0f} đ"
            ):
                return
            
            # Tạo hóa đơn
            count = 0
            errors = []
            
            import random
            
            for contract in active_contracts:
                try:
                    contract_id = contract[0]
                    student_id = contract[1]
                    room_fee = float(contract[5])  # MonthlyFee
                    
                    # Tính phí điện/nước ngẫu nhiên (giả lập số đếm)
                    electricity_kwh = random.randint(40, 100)
                    water_m3 = random.randint(5, 15)
                    
                    electricity_fee = electricity_kwh * elec_price
                    water_fee = water_m3 * water_price
                    
                    # Tạo hóa đơn
                    success = self.invoice_dao.create_invoice(
                        contract_id, student_id, month_str,
                        room_fee, electricity_fee, water_fee,
                        internet_fee, service_fee
                    )
                    
                    if success:
                        count += 1
                    else:
                        errors.append(f"Hợp đồng {contract_id}")
                
                except Exception as e:
                    errors.append(f"Hợp đồng {contract[0]}: {str(e)}")
            
            # Hiển thị kết quả
            result_msg = f"Đã tạo {count}/{len(active_contracts)} hóa đơn cho tháng {month_str}!"
            
            if errors:
                result_msg += f"\n\nCó {len(errors)} lỗi:\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    result_msg += f"\n... và {len(errors)-5} lỗi khác"
                messagebox.showwarning("Hoàn thành với lỗi", result_msg)
            else:
                messagebox.showinfo("Thành công", result_msg)
            
            self.callback()
            self.dialog.destroy()
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể tạo hóa đơn:\n{str(e)}")
            import traceback
            traceback.print_exc()


# ============================================
# Dialog thanh toán
# ============================================

class PaymentDialog:
    def __init__(self, parent, invoice, invoice_dao, callback):
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Thanh toán hóa đơn")
        self.dialog.geometry("400x300")
        self.dialog.grab_set()
        
        self.invoice = invoice
        self.invoice_dao = invoice_dao
        self.callback = callback
        
        self.create_widgets()
    
    def create_widgets(self):
        """Tạo form"""
        tk.Label(
            self.dialog,
            text="THANH TOÁN HÓA ĐƠN",
            font=('Arial', 14, 'bold'),
            fg='#2196F3'
        ).pack(pady=15)
        
        # Thông tin
        info_frame = tk.Frame(self.dialog)
        info_frame.pack(padx=20, pady=10)
        
        tk.Label(
            info_frame,
            text=f"Mã HĐ: {self.invoice[0]} - {self.invoice[16]} - Phòng {self.invoice[17]}",
            font=('Arial', 10, 'bold')
        ).pack()
        
        tk.Label(
            info_frame,
            text=f"Tổng tiền: {self.invoice[9]:,.0f} đ",
            font=('Arial', 10)
        ).pack()
        
        tk.Label(
            info_frame,
            text=f"Còn nợ: {self.invoice[11]:,.0f} đ",
            font=('Arial', 11, 'bold'),
            fg='red'
        ).pack()
        
        # Form
        form_frame = tk.Frame(self.dialog)
        form_frame.pack(padx=20, pady=10)
        
        tk.Label(form_frame, text="Số tiền thanh toán (*):").grid(row=0, column=0, sticky='w', pady=10)
        self.amount_entry = tk.Entry(form_frame, width=25)
        self.amount_entry.insert(0, self.invoice[11])
        self.amount_entry.grid(row=0, column=1, pady=10)
        
        tk.Label(form_frame, text="Hình thức:").grid(row=1, column=0, sticky='w', pady=10)
        self.method_combo = ttk.Combobox(form_frame, width=23, state='readonly')
        self.method_combo['values'] = ['Tiền mặt', 'Chuyển khoản', 'Thẻ']
        self.method_combo.current(0)
        self.method_combo.grid(row=1, column=1, pady=10)
        
        # Buttons
        btn_frame = tk.Frame(self.dialog)
        btn_frame.pack(pady=20)
        
        tk.Button(
            btn_frame,
            text="✅ Xác nhận",
            bg='#4CAF50',
            fg='white',
            width=12,
            command=self.pay
        ).pack(side='left', padx=5)
        
        tk.Button(
            btn_frame,
            text="❌ Hủy",
            bg='#f44336',
            fg='white',
            width=12,
            command=self.dialog.destroy
        ).pack(side='left', padx=5)
    
    def pay(self):
        """Thanh toán"""
        try:
            amount = Decimal(self.amount_entry.get())
            
            if amount <= 0:
                messagebox.showwarning("Cảnh báo", "Số tiền phải lớn hơn 0!")
                return
            
            if amount > self.invoice[11]:
                if not messagebox.askyesno("Xác nhận", 
                    f"Số tiền thanh toán ({amount:,.0f} đ) lớn hơn số nợ ({self.invoice[11]:,.0f} đ). Tiếp tục?"):
                    return
            
            method = self.method_combo.get()
            
            if self.invoice_dao.record_payment(self.invoice[0], amount, method, ''):
                messagebox.showinfo("Thành công", "Đã ghi nhận thanh toán!")
                self.callback()
                self.dialog.destroy()
            else:
                messagebox.showerror("Lỗi", "Không thể ghi nhận thanh toán!")
                
        except ValueError:
            messagebox.showwarning("Cảnh báo", "Số tiền không hợp lệ!")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi: {e}")