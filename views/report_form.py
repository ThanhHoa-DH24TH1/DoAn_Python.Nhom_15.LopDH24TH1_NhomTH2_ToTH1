import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from dao.student_dao import StudentDAO
from dao.room_dao import RoomDAO
from dao.invoice_dao import InvoiceDAO
import openpyxl
from openpyxl.styles import Font

class ReportForm:
    def __init__(self, parent):
        self.window = tk.Toplevel(parent)
        self.window.title("Báo cáo & Thống kê")
        self.window.geometry("1200x700")
        self.window.state('zoomed')
        
        self.student_dao = StudentDAO()
        self.room_dao = RoomDAO()
        self.invoice_dao = InvoiceDAO()
        
        self.create_widgets()
    
    def create_widgets(self):
        """Tạo giao diện"""
        # Title
        title_frame = tk.Frame(self.window, bg="#00F2FF", height=60)
        title_frame.pack(fill='x')
        title_frame.pack_propagate(False)
        
        tk.Label(
            title_frame,
            text="BÁO CÁO & THỐNG KÊ",
            font=('Arial', 16, 'bold'),
            bg="#00F2FF",
            fg='white'
        ).pack(pady=15)
        
        # Notebook (Tabs)
        self.notebook = ttk.Notebook(self.window)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Tab 1: Thống kê sinh viên
        self.create_student_stats_tab()
        
        # Tab 2: Báo cáo doanh thu
        self.create_revenue_tab()
        
        # Tab 3: Tình trạng phòng
        self.create_room_stats_tab()
    
    def create_student_stats_tab(self):
        """Tab thống kê sinh viên"""
        tab = tk.Frame(self.notebook)
        self.notebook.add(tab, text="📊 Thống kê Sinh viên")
        
        # Filter frame
        filter_frame = tk.Frame(tab, bg='white')
        filter_frame.pack(fill='x', padx=10, pady=10)
        
        tk.Label(filter_frame, text="Thống kê theo:", bg='white').pack(side='left', padx=5)
        self.student_type_combo = ttk.Combobox(filter_frame, width=15, state='readonly')
        self.student_type_combo['values'] = ['Khoa', 'Lớp', 'Tòa nhà']
        self.student_type_combo.current(0)
        self.student_type_combo.pack(side='left', padx=5)
        
        tk.Button(
            filter_frame,
            text="🔍 Xem thống kê",
            bg='#2196F3',
            fg='white',
            command=self.show_student_stats
        ).pack(side='left', padx=5)
        
        
        
        # Content frame
        content_frame = tk.Frame(tab)
        content_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Left: Chart
        chart_frame = tk.LabelFrame(content_frame, text="Biểu đồ", font=('Arial', 10, 'bold'))
        chart_frame.pack(side='left', fill='both', expand=True)
        
        self.student_figure = Figure(figsize=(6, 5))
        self.student_canvas = FigureCanvasTkAgg(self.student_figure, chart_frame)
        self.student_canvas.get_tk_widget().pack(fill='both', expand=True, padx=10, pady=10)
        
        # Right: Table
        table_frame = tk.LabelFrame(content_frame, text="Số liệu chi tiết", font=('Arial', 10, 'bold'), width=400)
        table_frame.pack(side='right', fill='both', padx=(10, 0))
        table_frame.pack_propagate(False)
        
        columns = ('Nhóm', 'Số lượng', 'Tỷ lệ %')
        self.student_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)
        
        for col in columns:
            self.student_tree.heading(col, text=col)
            self.student_tree.column(col, width=120, anchor='center')
        
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.student_tree.yview)
        self.student_tree.configure(yscrollcommand=vsb.set)
        
        self.student_tree.pack(side='left', fill='both', expand=True, padx=(10, 0), pady=10)
        vsb.pack(side='right', fill='y', pady=10)
        
        # Load initial data
        self.show_student_stats()
    
    def show_student_stats(self):
        """Hiển thị thống kê sinh viên"""
        stat_type = self.student_type_combo.get()
        
        # Clear tree
        for item in self.student_tree.get_children():
            self.student_tree.delete(item)
        
        # Get data
        students = self.student_dao.get_all_students()
        
        if stat_type == 'Khoa':
            stats = {}
            for s in students:
                faculty = s[9] or 'Chưa xác định'
                stats[faculty] = stats.get(faculty, 0) + 1
        elif stat_type == 'Lớp':
            stats = {}
            for s in students:
                class_name = s[11] or 'Chưa xác định'
                stats[class_name] = stats.get(class_name, 0) + 1
        else:  # Tòa nhà
            stats = {}
            for s in students:
                building = s[14].split(' - ')[0] if s[14] else 'Chưa có phòng'
                if building and building != 'Chưa có phòng':
                    building = building[0]  # Lấy ký tự đầu (A, B, C...)
                stats[building] = stats.get(building, 0) + 1
        
        total = sum(stats.values())
        
        # Update tree
        for name, count in sorted(stats.items()):
            percent = (count / total * 100) if total > 0 else 0
            self.student_tree.insert('', 'end', values=(name, count, f"{percent:.1f}%"))
        
        # Draw chart
        self.student_figure.clear()
        ax = self.student_figure.add_subplot(111)
        
        labels = list(stats.keys())
        sizes = list(stats.values())
        colors = ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#FF9F40']
        
        ax.pie(sizes, labels=labels, autopct='%1.1f%%', colors=colors[:len(labels)], startangle=90)
        ax.set_title(f'Thống kê sinh viên theo {stat_type}', fontsize=12, fontweight='bold')
        
        self.student_canvas.draw()
    
    
    
    def create_revenue_tab(self):
        """Tab báo cáo doanh thu"""
        tab = tk.Frame(self.notebook)
        self.notebook.add(tab, text="💰 Báo cáo Doanh thu")
        
        # Filter frame
        filter_frame = tk.Frame(tab, bg='white')
        filter_frame.pack(fill='x', padx=10, pady=10)
        
        tk.Label(filter_frame, text="Từ tháng:", bg='white').pack(side='left', padx=5)
        self.from_month_combo = ttk.Combobox(filter_frame, width=10, state='readonly')
        self.from_month_combo['values'] = self.get_months()
        self.from_month_combo.current(6)
        self.from_month_combo.pack(side='left', padx=5)
        
        tk.Label(filter_frame, text="Đến tháng:", bg='white').pack(side='left', padx=5)
        self.to_month_combo = ttk.Combobox(filter_frame, width=10, state='readonly')
        self.to_month_combo['values'] = self.get_months()
        self.to_month_combo.current(12)
        self.to_month_combo.pack(side='left', padx=5)
        
        tk.Button(
            filter_frame,
            text="🔍 Xem báo cáo",
            bg='#2196F3',
            fg='white',
            command=self.show_revenue_report
        ).pack(side='left', padx=5)
        
        tk.Button(
            filter_frame,
            text="📄 Xuất Excel",
            bg='#4CAF50',
            fg='white',
            command=self.export_revenue_report
        ).pack(side='left', padx=5)
        
        # Content frame
        content_frame = tk.Frame(tab)
        content_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Top: Chart
        chart_frame = tk.LabelFrame(content_frame, text="Biểu đồ doanh thu", font=('Arial', 10, 'bold'))
        chart_frame.pack(fill='both', expand=True)
        
        self.revenue_figure = Figure(figsize=(10, 4))
        self.revenue_canvas = FigureCanvasTkAgg(self.revenue_figure, chart_frame)
        self.revenue_canvas.get_tk_widget().pack(fill='both', expand=True, padx=10, pady=10)
        
        # Bottom: Table
        table_frame = tk.Frame(content_frame)
        table_frame.pack(fill='x', pady=(10, 0))
        
        columns = ('Tháng', 'Tiền phòng', 'Tiền điện', 'Tiền nước', 'Dịch vụ', 'Tổng cộng')
        self.revenue_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=8)
        
        widths = [100, 150, 150, 150, 150, 150]
        for col, width in zip(columns, widths):
            self.revenue_tree.heading(col, text=col)
            self.revenue_tree.column(col, width=width, anchor='center')
        
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.revenue_tree.yview)
        self.revenue_tree.configure(yscrollcommand=vsb.set)
        
        self.revenue_tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')
        
        # Total label
        self.total_revenue_label = tk.Label(
            tab,
            text="",
            font=('Arial', 12, 'bold'),
            fg='#4CAF50'
        )
        self.total_revenue_label.pack(pady=10)
        
        # Load initial data
        self.show_revenue_report()
    
    def get_months(self):
        """Lấy danh sách tháng"""
        from datetime import datetime
        months = []
        for i in range(12, -1, -1):
            date = datetime.now()
            month = date.month - i
            year = date.year
            if month <= 0:
                month += 12
                year -= 1
            months.append(f"{year}-{month:02d}")
        return months
    
    def show_revenue_report(self):
        """Hiển thị báo cáo doanh thu"""
        # Clear tree
        for item in self.revenue_tree.get_children():
            self.revenue_tree.delete(item)
        
        from_month = self.from_month_combo.get()
        to_month = self.to_month_combo.get()
        
        # Get all months in range
        all_months = self.get_months()
        start_idx = all_months.index(from_month)
        end_idx = all_months.index(to_month)
        
        if start_idx > end_idx:
            messagebox.showwarning("Cảnh báo", "Tháng bắt đầu phải trước tháng kết thúc!")
            return
        
        selected_months = all_months[start_idx:end_idx+1]
        
        # Collect data
        month_data = {}
        total_all = 0
        
        for month in selected_months:
            invoices = self.invoice_dao.get_all_invoices(month=month)
            
            room_fee = sum(inv[4] for inv in invoices)
            elec_fee = sum(inv[5] for inv in invoices)
            water_fee = sum(inv[6] for inv in invoices)
            service_fee = sum(inv[7] + inv[8] for inv in invoices)
            total = room_fee + elec_fee + water_fee + service_fee
            
            month_data[month] = {
                'room': room_fee,
                'elec': elec_fee,
                'water': water_fee,
                'service': service_fee,
                'total': total
            }
            total_all += total
            
            # Add to tree
            self.revenue_tree.insert('', 'end', values=(
                month,
                f"{room_fee:,.0f}",
                f"{elec_fee:,.0f}",
                f"{water_fee:,.0f}",
                f"{service_fee:,.0f}",
                f"{total:,.0f}"
            ))
        
        # Update total label
        self.total_revenue_label.config(
            text=f"💰 TỔNG DOANH THU: {total_all:,.0f} VNĐ"
        )
        
        # Draw chart
        self.revenue_figure.clear()
        ax = self.revenue_figure.add_subplot(111)
        
        months = list(month_data.keys())
        totals = [month_data[m]['total'] for m in months]
        
        ax.plot(months, totals, marker='o', linewidth=2, markersize=8, color='#4CAF50')
        ax.set_xlabel('Tháng', fontsize=10)
        ax.set_ylabel('Doanh thu (VNĐ)', fontsize=10)
        ax.set_title('Biểu đồ doanh thu theo tháng', fontsize=12, fontweight='bold')
        ax.grid(True, alpha=0.3)
        ax.tick_params(axis='x', rotation=45)
        
        self.revenue_figure.tight_layout()
        self.revenue_canvas.draw()
    
    def export_revenue_report(self):
        """Xuất Excel"""
        # 1. Hỏi người dùng muốn lưu file ở đâu
        file_path = filedialog.asksaveasfilename(
            title="Lưu file Báo cáo Doanh thu",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        # 2. Nếu người dùng không chọn (nhấn Cancel) thì dừng lại
        if not file_path:
            return

        try:
            # 3. Tạo file Excel mới
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BaoCaoDoanhThu"
            
            # 4. Lấy tiêu đề cột từ Treeview
            headers = list(self.revenue_tree['columns'])
            
            # 5. Ghi tiêu đề (in đậm)
            ws.append(headers)
            header_font = Font(bold=True)
            for cell in ws[1]: # ws[1] là hàng đầu tiên
                cell.font = header_font
            
            # 6. Ghi dữ liệu từng hàng từ Treeview
            for item_id in self.revenue_tree.get_children():
                row_values = self.revenue_tree.item(item_id)['values']
                ws.append(list(row_values))

            # 7. Thêm dòng tổng cộng ở cuối
            ws.append([]) # Thêm một hàng trống
            
            total_text = self.total_revenue_label.cget("text") # Lấy text từ Label
            total_cell = ws.cell(row=ws.max_row + 1, column=1) # Ô đầu tiên của hàng mới
            total_cell.value = total_text
            total_cell.font = Font(bold=True, size=14, color="008000") # In đậm, cỡ 14, màu xanh
                
            # 8. (Tùy chọn) Tự động điều chỉnh độ rộng cột
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Lấy tên cột (A, B, C...)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # 9. Lưu file 💾
            wb.save(file_path)
            
            messagebox.showinfo("Thành công", f"Đã xuất báo cáo ra file:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu file Excel:\n{e}")
    
    def create_room_stats_tab(self):
        """Tab tình trạng phòng"""
        tab = tk.Frame(self.notebook)
        self.notebook.add(tab, text="🏢 Tình trạng Phòng")
        
        # Filter frame
        filter_frame = tk.Frame(tab, bg='white')
        filter_frame.pack(fill='x', padx=10, pady=10)
        
        tk.Label(filter_frame, text="Tòa nhà:", bg='white').pack(side='left', padx=5)
        self.room_building_combo = ttk.Combobox(filter_frame, width=10, state='readonly')
        self.room_building_combo.pack(side='left', padx=5)
        
        tk.Button(
            filter_frame,
            text="🔍 Xem thống kê",
            bg='#2196F3',
            fg='white',
            command=self.show_room_stats
        ).pack(side='left', padx=5)
        
        # Stats cards frame
        stats_frame = tk.Frame(tab)
        stats_frame.pack(fill='x', padx=10, pady=10)
        
        self.room_total_label = self.create_stat_label(stats_frame, "Tổng số phòng", "0", '#2196F3', 0)
        self.room_occupied_label = self.create_stat_label(stats_frame, "Đang sử dụng", "0", '#4CAF50', 1)
        self.room_empty_label = self.create_stat_label(stats_frame, "Phòng trống", "0", '#FF9800', 2)
        self.room_rate_label = self.create_stat_label(stats_frame, "Tỷ lệ lấp đầy", "0%", '#9C27B0', 3)
        
        # Content frame
        content_frame = tk.Frame(tab)
        content_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Left: Chart
        chart_frame = tk.LabelFrame(content_frame, text="Biểu đồ", font=('Arial', 10, 'bold'))
        chart_frame.pack(side='left', fill='both', expand=True)
        
        self.room_figure = Figure(figsize=(6, 5))
        self.room_canvas = FigureCanvasTkAgg(self.room_figure, chart_frame)
        self.room_canvas.get_tk_widget().pack(fill='both', expand=True, padx=10, pady=10)
        
        # Right: Table
        table_frame = tk.LabelFrame(content_frame, text="Chi tiết theo tầng", font=('Arial', 10, 'bold'), width=400)
        table_frame.pack(side='right', fill='both', padx=(10, 0))
        table_frame.pack_propagate(False)
        
        columns = ('Tầng', 'Tổng phòng', 'Đang ở', 'Trống')
        self.room_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)
        
        for col in columns:
            self.room_tree.heading(col, text=col)
            self.room_tree.column(col, width=90, anchor='center')
        
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.room_tree.yview)
        self.room_tree.configure(yscrollcommand=vsb.set)
        
        self.room_tree.pack(side='left', fill='both', expand=True, padx=(10, 0), pady=10)
        vsb.pack(side='right', fill='y', pady=10)
        
        # Load buildings
        buildings = self.room_dao.get_buildings()
        self.room_building_combo['values'] = ['Tất cả'] + buildings
        self.room_building_combo.current(0)
        
        # Load initial data
        self.show_room_stats()
    
    def create_stat_label(self, parent, title, value, color, col):
        """Tạo label thống kê"""
        frame = tk.Frame(parent, bg=color, relief='raised', bd=2)
        frame.grid(row=0, column=col, padx=10, pady=10, sticky='nsew')
        parent.grid_columnconfigure(col, weight=1)
        
        tk.Label(
            frame,
            text=title,
            font=('Arial', 10),
            bg=color,
            fg='white'
        ).pack(pady=(15, 5))
        
        label = tk.Label(
            frame,
            text=value,
            font=('Arial', 20, 'bold'),
            bg=color,
            fg='white'
        )
        label.pack(pady=(5, 15))
        
        return label
    
    def show_room_stats(self):
        """Hiển thị thống kê phòng"""
        building = self.room_building_combo.get()
        building = None if building == 'Tất cả' else building
        
        rooms = self.room_dao.search_rooms(building=building)
        
        # Clear tree
        for item in self.room_tree.get_children():
            self.room_tree.delete(item)
        
        # Calculate stats
        total_rooms = len(rooms)
        occupied = sum(1 for r in rooms if r[6] > 0)  # CurrentOccupancy > 0
        empty = sum(1 for r in rooms if r[6] == 0)
        rate = (occupied / total_rooms * 100) if total_rooms > 0 else 0
        
        # Update labels
        self.room_total_label.config(text=str(total_rooms))
        self.room_occupied_label.config(text=str(occupied))
        self.room_empty_label.config(text=str(empty))
        self.room_rate_label.config(text=f"{rate:.1f}%")
        
        # Stats by floor
        floor_stats = {}
        for r in rooms:
            floor = r[3]
            if floor not in floor_stats:
                floor_stats[floor] = {'total': 0, 'occupied': 0, 'empty': 0}
            floor_stats[floor]['total'] += 1
            if r[6] > 0:
                floor_stats[floor]['occupied'] += 1
            else:
                floor_stats[floor]['empty'] += 1
        
        # Update tree
        for floor in sorted(floor_stats.keys()):
            stats = floor_stats[floor]
            self.room_tree.insert('', 'end', values=(
                f"Tầng {floor}",
                stats['total'],
                stats['occupied'],
                stats['empty']
            ))
        
        # Draw chart
        self.room_figure.clear()
        ax = self.room_figure.add_subplot(111)
        
        labels = ['Đang sử dụng', 'Phòng trống']
        sizes = [occupied, empty]
        colors = ['#4CAF50', '#FF9800']
        explode = (0.1, 0)
        
        ax.pie(sizes, explode=explode, labels=labels, autopct='%1.1f%%',
               colors=colors, startangle=90, shadow=True)
        ax.set_title('Tình trạng phòng', fontsize=12, fontweight='bold')
        
        self.room_canvas.draw()
